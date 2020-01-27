# Python program to read an excel file

# import openpyxl module
import openpyxl

# Give the location of the file
path = "C:\\Users\\intel\\Desktop\\Medicine Preciption\\Syn.xlsm"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

max_col = sheet_obj.max_column

# Will print a particular row value
for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row=2, column=i)
    print(cell_obj.value, end=" ")
